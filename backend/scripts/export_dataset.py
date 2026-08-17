#!/usr/bin/env python
"""Export the current database into a readable synthetic demo workbook."""
from __future__ import annotations

from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(ROOT))

from app.db.models import ActivityEvent, Applicant, Application, ClientPreference, Conversation, Feedback, Interaction, Property, SavedProperty, Viewing, ViewingRequest
from app.db.session import SessionLocal
from app.intelligence.matching.engine import PropertyMatchingEngine


def rows(db, model):
    columns = [column.name for column in model.__table__.columns]
    return [{column: getattr(row, column) for column in columns} for row in db.scalars(select(model)).all()]


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    output = root / "exports" / "property_intelligence_demo_dataset.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    db = SessionLocal()
    try:
        sheets = {
            "Applicants": rows(db, Applicant),
            "Properties": rows(db, Property),
            "Preferences": rows(db, ClientPreference),
            "Conversations": rows(db, Conversation),
            "Interactions": rows(db, Interaction),
            "Viewings": rows(db, Viewing),
            "Feedback": rows(db, Feedback),
            "Applications": rows(db, Application),
            "Saved Properties": rows(db, SavedProperty),
            "Activity Events": rows(db, ActivityEvent),
            "Viewing Requests": rows(db, ViewingRequest),
        }
        sarah = db.get(Applicant, "A-DEMO-SARAH")
        matches = PropertyMatchingEngine(db).match(sarah, 20) if sarah else []
        sheets["Matches"] = [{"applicant_id": sarah.applicant_id, "property_id": match.property.property_id, "match_score": match.match_score, "positive_reasons": " | ".join(match.explanation.positives), "negative_reasons": " | ".join(match.explanation.negatives)} for match in matches]
        overview = [{"Field": "Dataset", "Value": "Synthetic demonstration dataset"}, {"Field": "Purpose", "Value": "CTO demo of property matching, workflow and grounded RAG"}, {"Field": "Generated from", "Value": "Current SQLAlchemy database"}]
        sheets = {"Overview": overview, **sheets}
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for name, data in sheets.items():
                frame = pd.DataFrame(data)
                if frame.empty:
                    frame = pd.DataFrame({"No records": []})
                frame.to_excel(writer, sheet_name=name[:31], index=False)
        workbook = load_workbook(output)
        header_fill = PatternFill("solid", fgColor="173D31")
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
            for column in sheet.columns:
                width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 48)
                sheet.column_dimensions[column[0].column_letter].width = width
        workbook.save(output)
        print(f"Exported {len(sheets)} sheets to {output}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
